# SPDX-FileCopyrightText: 2024-present U.N. Owen <void@some.where>
#
# SPDX-License-Identifier: MIT

#!/usr/bin/env python3
# PYTHON_ARGCOMPLETE_OK
import argparse
import logging
import math
import os
import socket
import sys
import time

import argcomplete
import openpyxl
import openpyxl.cell
import pandas
from pybatfish.client.session import Session
from pybatfish.datamodel.flow import (
    HeaderConstraints,
    MatchTcpFlags,
    PathConstraints,
    TcpFlags,
)

import bfqry.cli.parser as parser
import bfqry.cli.common as c
import bfqry.cli.util as util

time_start = time.perf_counter()


def bireach(
    bf: Session, action: str, args: argparse.ArgumentParser
) -> pandas.DataFrame:
    headers = get_headers(args)
    location = get_start_location(args)
    data = (
        bf.q.bidirectionalReachability(
            pathConstraints=PathConstraints(startLocation=location),
            headers=headers,
            returnFlowType=action.upper(),
        )
        .answer()
        .frame()
    )
    return data


def calc_speed(speed) -> str:
    if speed is None:
        return "(None)"
    if 0 <= speed and speed < 1000:
        return str(math.floor(speed))
    if 1000 <= speed and speed < 1000000:
        return str(math.floor(speed / 1000)) + "K"
    if 1000000 <= speed and speed < 1000000000:
        return str(math.floor(speed / 1000000)) + "M"
    if 1000000000 <= speed and speed < 1000000000000:
        return str(math.floor(speed / 1000000000)) + "G"
    if 1000000000000 <= speed and speed < 1000000000000000:
        return str(math.floor(speed / 1000000000000)) + "T"
    return "?"


def cmd_aclsearch(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Node", "Action", "Filter_Name"]
    headers = get_headers(args)
    data1 = (
        bf.q.searchFilters(
            nodes=get_query(args.node), filters=get_query(args.filter), headers=headers
        )
        .answer()
        .frame()
    ).sort_values(by=sort1)
    lines = []
    _total = len(data1)
    for _index, row in enumerate(data1.itertuples()):
        _node = row[1]
        _filter_name = row[2]
        _flow = row[3]
        _action = row[4]
        _line_content = row[5]
        _trace = row[6]
        lines.append(f"{c.SEPARATOR1} Flow {_index + 1}/{_total} ({_action})")
        lines.append(f"Node         : {_node}")
        lines.append(f"Filter Name  : {_filter_name}")
        lines.append(f"Flow         : {_flow}")
        lines.append(f"Action       : {_action}")
        lines.append(f"Line Content : {_line_content}")
        lines.append(f"Trace        : {_trace}")
    show_result_lines(args, "\n".join(lines))


def cmd_acltest(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Node", "Action", "Filter_Name"]
    headers = get_headers(args)
    data1 = (
        bf.q.testFilters(
            nodes=get_query(args.node), filters=get_query(args.filter), headers=headers
        )
        .answer()
        .frame()
    ).sort_values(by=sort1)
    lines = []
    _total = len(data1)
    for _index, row in enumerate(data1.itertuples()):
        _node = row[1]
        _filter_name = row[2]
        _flow = row[3]
        _action = row[4]
        _line_content = row[5]
        _trace = row[6]
        lines.append(f"{c.SEPARATOR1} Flow {_index + 1}/{_total} ({_action})")
        lines.append(f"Node         : {_node}")
        lines.append(f"Filter Name  : {_filter_name}")
        lines.append(f"Flow         : {_flow}")
        lines.append(f"Action       : {_action}")
        lines.append(f"Line Content : {_line_content}")
        lines.append(f"Trace        : {_trace}")
    show_result_lines(args, "\n".join(lines))


def cmd_aclunreach(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Sources", "Unreachable_Line"]
    sort2 = ["Sources", "Unreachable_Line_Action"]
    show1 = [
        "Sources",
        "Unreachable_Line",
        "Blocking_Lines",
        "Reason",
    ]
    data1 = (
        bf.q.filterLineReachability(
            nodes=get_query(args.node), filters=get_query(args.filter)
        )
        .answer()
        .frame()
    ).sort_values(by=sort2)
    lines = []
    _total = len(data1)
    if args.summary:
        show_result(args, data1[show1], sort1)
    else:
        for _index, row in enumerate(data1.itertuples()):
            _sources = row[1]
            _unreach_line = row[2]
            _unreach_action = row[3]
            _block_lines = row[4]
            _different_action = row[5]
            _reason = row[6]
            _info = row[7]
            lines.append(
                f"{c.SEPARATOR1} Flow {_index + 1}/{_total} ({_unreach_action})"
            )
            lines.append(f"Sources                 : {_sources}")
            lines.append(f"Unreachable Line        : {_unreach_line}")
            lines.append(f"Unreachable Line Action : {_unreach_action}")
            lines.append(f"Blocking Lines          : {_block_lines}")
            lines.append(f"Different Action        : {_different_action}")
            lines.append(f"Reason                  : {_reason}")
            lines.append(f"Additional Info         : {_info}")
        show_result_lines(args, "\n".join(lines))


def cmd_bgpedge(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Node", "AS_Number", "Remote_Node", "Remote_AS_Number"]
    show1 = [
        "Node",
        "IP",
        "AS_Number",
        "Remote_Node",
        "Remote_IP",
        "Remote_AS_Number",
    ]
    show2 = [
        "Node",
        "Interface",
        "IP",
        "AS_Number",
        "Remote_Node",
        "Remote_IP",
        "Remote_Interface",
        "Remote_AS_Number",
    ]
    data1 = bf.q.bgpEdges(nodes=get_query(args.node)).answer().frame()
    if args.detail:
        show_result(args, data1[show2], sort1)
    else:
        show_result(args, data1[show1], sort1)


def cmd_bgprib(bf: Session, args: argparse.ArgumentParser):
    # sort1 = ["Node", "Network", "Next_Hop", "Protocol"]
    show1 = [
        "Node",
        "Network",
        "Status",
        "Next_Hop",
        "Protocol",
        "AS_Path",
        "Metric",
        "Local_Pref",
        "Weight",
        "Tag",
    ]
    data1 = bf.q.bgpRib(nodes=get_query(args.node)).answer().frame()
    # show_result だとエラーになってしまう。
    # (datafarame の Sort 処理にバグがありそう？)
    # その為、show_result_lines を使う。
    # 下記だとエラーになってしまう… (バグ？)
    show_result_lines(args, data1[show1].to_string(index=False))


def cmd_bireach(bf: Session, args: argparse.ArgumentParser):
    total_flow = 0
    if args.success:
        actions = ["SUCCESS"]
    else:
        actions = ["SUCCESS", "FAILURE", "MULTIPATH_INCONSISTENT"]
    lines = []
    for action in actions:
        util.debug(
            f"Bi-directional Reachability Question ({action}) is being executed."
        )
        data1 = bireach(bf, action, args)
        # Flow count.
        debug_bidir_flow(data1, f"'{action}' Flow Count")
        total_flow += data1.Forward_Flow.count()
        # Forward Flows
        for i, flow in enumerate(data1.Forward_Flow):
            lines.extend(
                [f"{c.SEPARATOR1} {i + 1}/{len(data1.Forward_Flow)} {action} Flow(s)"]
            )
            lines.extend([f"{data1.Forward_Flow[i]}"])
            lines.extend([f"{data1.Reverse_Flow[i]}"])
            # Forward Trace
            ftotal = len(data1.Forward_Traces[i])
            for findex_trace, ftrace in enumerate(data1["Forward_Traces"][i]):
                if args.format != "flow":
                    lines.extend([f"{c.SEPARATOR2} {ftotal} Forward Trace(s)"])
                fseparator = f"{c.SEPARATOR3} {findex_trace + 1}/{ftotal} Forward Trace"
                lines.extend(get_formatted_result(args, fseparator, ftrace))

            # Reverse Flows
            if data1.Reverse_Flow[i] is None:
                continue
            total_flow += data1.Reverse_Flow.count()
            # Reverse Trace
            rtotal = len(data1.Reverse_Traces[i])
            for rindex_trace, rtrace in enumerate(data1["Reverse_Traces"][i]):
                if args.format != "flow":
                    lines.extend([f"{c.SEPARATOR2} {rtotal} Reverse Trace(s)"])
                rseparator = f"{c.SEPARATOR3} {rindex_trace + 1}/{rtotal} Reverse Trace"
                lines.extend(get_formatted_result(args, rseparator, rtrace))
    if total_flow == 0:
        cmd_reach(bf, args)
    else:
        show_result_lines(args, "\n".join(lines))


def cmd_bitrace(bf: Session, args: argparse.ArgumentParser):
    headers = get_headers(args)
    location = get_start_location(args)
    data1 = (
        bf.q.bidirectionalTraceroute(startLocation=location, headers=headers)
        .answer()
        .frame()
    )
    lines = []

    # Flow count.
    debug_bidir_flow(data1, "Flow Count")

    # Flows
    for i, flow in enumerate(data1.Forward_Flow):
        lines.extend([f"{c.SEPARATOR1} {i + 1}/{len(data1.Forward_Flow)} Flow(s)"])
        lines.extend([f"{data1.Forward_Flow[i]}"])
        lines.extend([f"{data1.Reverse_Flow[i]}"])

        # Forward Trace
        ftotal = len(data1.Forward_Traces[i])
        for j, ftrace in enumerate(data1["Forward_Traces"][i]):
            if args.format != "flow":
                lines.extend([f"{c.SEPARATOR2} {ftotal} Forward Trace(s)"])
            fseparator = f"{c.SEPARATOR3} {j + 1}/{ftotal} Forward Trace"
            lines.extend(get_formatted_result(args, fseparator, ftrace))

        # Reverse Trace
        rtotal = len(data1.Reverse_Traces[i])
        for k, rtrace in enumerate(data1["Reverse_Traces"][i]):
            if args.format != "flow":
                lines.extend([f"{c.SEPARATOR2} {rtotal} Reverse Trace(s)"])
            rseparator = f"{c.SEPARATOR3} {k + 1}/{rtotal} Reverse Trace"
            lines.extend(get_formatted_result(args, rseparator, rtrace))
    show_result_lines(args, "\n".join(lines))


def cmd_eigrpedge(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Interface", "Remote_Interface"]
    show1 = ["Interface", "Remote_Interface"]
    data1 = bf.q.eigrpEdges(nodes=get_query(args.node)).answer().frame()
    show_result(args, data1[show1], sort1)


def cmd_hsrp(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Interface", "Group_Id"]
    show1 = [
        "Interface",
        "Group_Id",
        "Virtual_Addresses",
        "Source_Address",
        "Priority",
        "Preempt",
        "Active",
    ]
    data1 = bf.q.hsrpProperties(nodes=get_query(args.node)).answer().frame()
    if len(data1) != 0:
        data1["Virtual_Addresses"] = data1.apply(get_hsrp_vip, axis=1)
    if args.id is not None:
        data1 = data1[data1["Group_Id"] == args.id]
    if args.vip is not None:
        data1 = data1[data1["Virtual_Addresses"] == args.vip]
    if args.wopair:
        data1 = data1.drop_duplicates(
            subset=["Group_Id", "Virtual_Addresses"], keep=False
        )
    show_result(args, data1[show1], sort1)


def cmd_interface(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Interface"]
    show_acl = [
        "Interface",
        "Primary_Address",
        "Incoming_Filter_Name",
        "Outgoing_Filter_Name",
    ]
    show_all = [
        "Interface",
        "Access_VLAN",
        "Active",
        "Admin_Up",
        "All_Prefixes",
        "Allowed_VLANs",
        "Auto_State_VLAN",
        "Bandwidth",
        "Blacklisted",
        "Channel_Group",
        "Channel_Group_Members",
        "DHCP_Relay_Addresses",
        "Declared_Names",
        "Description",
        "Encapsulation_VLAN",
        "HSRP_Groups",
        "HSRP_Version",
        "Inactive_Reason",
        "Incoming_Filter_Name",
        "MLAG_ID",
        "MTU",
        "Native_VLAN",
        "Outgoing_Filter_Name",
        "PBR_Policy_Name",
        "Primary_Address",
        "Primary_Network",
        "Proxy_ARP",
        "Rip_Enabled",
        "Rip_Passive",
        "Spanning_Tree_Portfast",
        "Speed",
        "Switchport",
        "Switchport_Mode",
        "Switchport_Trunk_Encapsulation",
        "VRF",
        "VRRP_Groups",
        "Zone_Name",
    ]
    show_basic = [
        "Interface",
        "Primary_Address",
        "Description",
        "VRF",
    ]
    show_hsrp = [
        "Interface",
        "Primary_Address",
        "HSRP_Groups",
        "HSRP_Version",
    ]
    show_ip = [
        "Interface",
        "Primary_Address",
        "Primary_Network",
        "All_Prefixes",
        "VRF",
    ]
    show_portchannel = [
        "Interface",
        "Primary_Address",
        "Channel_Group",
        "Channel_Group_Members",
        "MLAG_ID",
    ]
    show_speed = [
        "Interface",
        "Primary_Address",
        "Speed",
        "Bandwidth",
    ]
    show_state = [
        "Interface",
        "Primary_Address",
        "Active",
        "Admin_Up",
        "Blacklisted",
        "Inactive_Reason",
    ]
    show_stp = [
        "Interface",
        "Primary_Address",
        "Switchport",
        "Spanning_Tree_Portfast",
    ]
    show_vlan = [
        "Interface",
        "Primary_Address",
        "Switchport",
        "Switchport_Mode",
        "Switchport_Trunk_Encapsulation",
        "Access_VLAN",
        "Allowed_VLANs",
        "Encapsulation_VLAN",
        "Native_VLAN",
    ]
    data1 = bf.q.interfaceProperties(nodes=get_query(args.node)).answer().frame()
    if not args.wounit:
        data1["Speed"] = data1.apply(get_speed, axis=1)
        data1["Bandwidth"] = data1.apply(get_bandwidth, axis=1)
    if args.column == "acl":
        show_result(args, data1[show_acl], sort1)
    elif args.column == "all":
        show_result(args, data1[show_all], sort1)
    elif args.column == "basic":
        show_result(args, data1[show_basic], sort1)
    elif args.column == "hsrp":
        show_result(args, data1[show_hsrp], sort1)
    elif args.column == "ip":
        show_result(args, data1[show_ip], sort1)
    elif args.column == "portchannel":
        show_result(args, data1[show_portchannel], sort1)
    elif args.column == "speed":
        show_result(args, data1[show_speed], sort1)
    elif args.column == "state":
        show_result(args, data1[show_state], sort1)
    elif args.column == "stp":
        show_result(args, data1[show_stp], sort1)
    elif args.column == "vlan":
        show_result(args, data1[show_vlan], sort1)


def cmd_ip(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Node", "VRF", "IP"]
    show1 = ["Node", "VRF", "IP", "Mask", "Interface", "Active"]
    data1 = (
        bf.q.ipOwners(ips=args.address, duplicatesOnly=args.duplicate).answer().frame()
    )

    if args.node is not None:
        data1 = data1[data1["Node"].isin(args.node)]
    if args.vrf is not None:
        data1 = data1[data1["VRF"] == args.vrf]
    show_result(args, data1[show1], sort1)


def cmd_l3edge(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Interface", "Remote_Interface", "IPs", "Remote_IPs"]
    show1 = ["Interface", "IPs", "Remote_Interface", "Remote_IPs"]
    data1 = (
        bf.q.layer3Edges(nodes=get_query(args.node), remoteNodes=get_query(args.rnode))
        .answer()
        .frame()
    )
    show_result(args, data1[show1], sort1)


def cmd_loop(bf: Session, args: argparse.ArgumentParser):
    data1 = bf.q.detectLoops().answer().frame()
    lines = []
    # Flow count.
    debug_flow(data1, "Flow Count")

    # Flows
    for i, flow in enumerate(data1.Flow):
        lines.extend([f"{c.SEPARATOR1} {i + 1}/{len(data1.Flow)} Flow"])
        lines.extend([f"{data1.Flow[i]}"])
        # Traces
        total = len(data1.Traces[i])
        for j, trace in enumerate(data1["Traces"][i]):
            separator = f"{c.SEPARATOR2} {j + 1}/{total} Trace"
            lines.extend(get_formatted_result(args, separator, trace))
    show_result_lines(args, "\n".join(lines))


def cmd_lpm(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Node", "Ip"]
    show1 = [
        "Node",
        "VRF",
        "Ip",
        "Network",
        "Num_Routes",
    ]
    data1 = bf.q.lpmRoutes(ip=args.address).answer().frame()
    if args.node is not None:
        data1 = data1[data1["Node"].isin(args.node)]
    if args.vrf is not None:
        data1 = data1[data1["VRF"] == args.vrf]
    show_result(args, data1[show1], sort1)


def cmd_node(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Node"]
    show_acl = [
        "Node",
        "Default_Inbound_Action",
        "Default_Cross_Zone_Action",
        "IP_Access_Lists",
        "IP6_Access_Lists",
    ]
    show_all = [
        "Node",
        "AS_Path_Access_Lists",
        "Authentication_Key_Chains",
        "Community_Match_Exprs",
        "Community_Set_Exprs",
        "Community_Set_Match_Exprs",
        "Community_Sets",
        "Configuration_Format",
        "DNS_Servers",
        "DNS_Source_Interface",
        "Default_Cross_Zone_Action",
        "Default_Inbound_Action",
        "Domain_Name",
        "Hostname",
        "IKE_Phase1_Keys",
        "IKE_Phase1_Policies",
        "IKE_Phase1_Proposals",
        "IP6_Access_Lists",
        "IP_Access_Lists",
        "IPsec_Peer_Configs",
        "IPsec_Phase2_Policies",
        "IPsec_Phase2_Proposals",
        "Interfaces",
        "Logging_Servers",
        "Logging_Source_Interface",
        "NTP_Servers",
        "NTP_Source_Interface",
        "PBR_Policies",
        "Route6_Filter_Lists",
        "Route_Filter_Lists",
        "Routing_Policies",
        "SNMP_Source_Interface",
        "SNMP_Trap_Servers",
        "TACACS_Servers",
        "TACACS_Source_Interface",
        "VRFs",
        "Zones",
    ]
    show_basic = [
        "Node",
        "Configuration_Format",
        "Interfaces",
        "VRFs",
        "Zones",
    ]
    show_bgp = [
        "Node",
        "AS_Path_Access_Lists",
        "Community_Match_Exprs",
        "Community_Set_Exprs",
        "Community_Set_Match_Exprs",
        "Community_Sets",
    ]
    show_dns = [
        "Node",
        "Domain_Name",
        "DNS_Servers",
        "DNS_Source_Interface",
    ]
    show_ipsec = [
        "Node",
        "IPsec_Peer_Configs",
        "IKE_Phase1_Keys",
        "IKE_Phase1_Policies",
        "IKE_Phase1_Proposals",
        "IPsec_Phase2_Policies",
        "IPsec_Phase2_Proposals",
    ]
    show_logging = [
        "Node",
        "Logging_Servers",
        "Logging_Source_Interface",
    ]
    show_ntp = [
        "Node",
        "NTP_Servers",
        "NTP_Source_Interface",
    ]
    show_route = [
        "Node",
        "Route_Filter_Lists",
        "Route6_Filter_Lists",
        "Routing_Policies",
        "PBR_Policies",
    ]
    show_snmp = [
        "Node",
        "SNMP_Source_Interface",
        "SNMP_Trap_Servers",
    ]
    show_tacacs = [
        "Node",
        "TACACS_Servers",
        "TACACS_Source_Interface",
    ]
    node = get_query(args.node)
    data1 = bf.q.nodeProperties(nodes=node).answer().frame()
    if args.column == "acl":
        show_result(args, data1[show_acl], sort1)
    elif args.column == "all":
        show_result(args, data1[show_all], sort1)
    elif args.column == "basic":
        show_result(args, data1[show_basic], sort1)
    elif args.column == "bgp":
        show_result(args, data1[show_bgp], sort1)
    elif args.column == "dns":
        show_result(args, data1[show_dns], sort1)
    elif args.column == "ipsec":
        show_result(args, data1[show_ipsec], sort1)
    elif args.column == "logging":
        show_result(args, data1[show_logging], sort1)
    elif args.column == "ntp":
        show_result(args, data1[show_ntp], sort1)
    elif args.column == "route":
        show_result(args, data1[show_route], sort1)
    elif args.column == "snmp":
        show_result(args, data1[show_snmp], sort1)
    elif args.column == "tacacs":
        show_result(args, data1[show_tacacs], sort1)


def cmd_ospfedge(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Interface", "Remote_Interface"]
    show1 = [
        "Interface",
        "Remote_Interface",
    ]
    data1 = bf.q.ospfEdges(nodes=get_query(args.node)).answer().frame()
    show_result(args, data1[show1], sort1)


def cmd_parse(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Nodes", "File_Name"]
    show1 = ["Nodes", "File_Name", "Status", "File_Format", "Version"]
    data1 = bf.q.fileParseStatus().answer().frame()

    data2 = pandas.DataFrame(data=None, columns=show1)
    for row in data1.itertuples():
        path = args.base + "/" + row[1]
        # Check OS version
        type = row[3]
        if type == "CISCO_IOS" or type == "CISCO_NX":
            version = get_os_version(path, "version ")
        elif type == "CISCO_ASA":
            version = get_os_version(path, "ASA Version ")
        else:
            version = "?"
        # Get node name from node name list.
        nodes = row[4]
        if len(nodes) == 0:
            hostname = ""
        else:
            hostname = nodes[0]
        # Concatenate the data.
        data2 = pandas.concat(
            [
                data2,
                pandas.DataFrame(
                    data=[[hostname, row[1], row[2], row[3], version]], columns=show1
                ),
            ]
        )
    if args.node is not None:
        data2 = data2[data2["Nodes"].isin(args.node)]
    show_result(args, data2[show1], sort1)


def cmd_prefix(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Node", "VRF", "Peer", "Action", "Prefix"]
    show1 = ["Node", "VRF", "Peer", "Action", "Prefix"]
    if args.prefix is None:
        data1 = (
            bf.q.prefixTracer(
                nodes=get_query(args.node),
            )
            .answer()
            .frame()
        )
    else:
        data1 = (
            bf.q.prefixTracer(
                nodes=get_query(args.node),
                prefix=args.prefix,
            )
            .answer()
            .frame()
        )
    show_result(args, data1[show1], sort1)


def cmd_reach(bf: Session, args: argparse.ArgumentParser):
    lines = []
    for action in ["SUCCESS", "FAILURE"]:
        data1 = reach(bf, action, args)
        # Flow count.
        debug_flow(data1, f"'{action}' Flow Count")
        # Flows
        for i, flow in enumerate(data1.Flow):
            lines.extend([f"{c.SEPARATOR1} {i + 1}/{len(data1.Flow)} {action} Flow"])
            lines.extend([f"{data1.Flow[i]}"])
            # Traces
            total = len(data1.Traces[i])
            for j, trace in enumerate(data1["Traces"][i]):
                separator = f"{c.SEPARATOR2} {j + 1}/{total} Trace"
                lines.extend(get_formatted_result(args, separator, trace))
    show_result_lines(args, "\n".join(lines))


def cmd_route(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Node", "Network", "Protocol"]
    show1 = [
        "Node",
        "Network",
        "Next_Hop_IP",
        "Next_Hop_Interface",
        "Protocol",
        "Metric",
        "Admin_Distance",
        "VRF",
        "Tag",
    ]
    show2 = [
        "Node",
        "Network",
        "Next_Hop_IP",
        "Protocol",
        "Metric",
        "Admin_Distance",
    ]

    # Routing protocol
    protocol = get_query(args.protocol)

    # Match parameter
    if args.match == "exact":
        match = "EXACT"
    elif args.match == "longest":
        match = "LONGEST_PREFIX_MATCH"
    elif args.match == "longer":
        match = "LONGER_PREFIXES"
    elif args.match == "shorter":
        match = "SHORTER_PREFIXES"

    if args.address is None:
        data1 = bf.q.routes(protocols=protocol, prefixMatchType=match).answer().frame()
    else:
        if is_ipaddr(args.address):
            address = args.address + "/32"
        else:
            address = args.address
        data1 = (
            bf.q.routes(network=address, protocols=protocol, prefixMatchType=match)
            .answer()
            .frame()
        )

    if args.node is not None:
        data1 = data1[data1["Node"].isin(args.node)]
    if args.vrf is not None:
        data1 = data1[data1["VRF"] == args.vrf]
    if args.detail:
        show_result(args, data1[show1], sort1)
    else:
        show_result(args, data1[show2], sort1)


def cmd_trace(bf: Session, args: argparse.ArgumentParser):
    headers = get_headers(args)
    location = get_start_location(args)
    data1 = bf.q.traceroute(startLocation=location, headers=headers).answer().frame()
    lines = []

    # Flow count.
    debug_flow(data1, "Flow Count")

    # Flows
    for i, flow in enumerate(data1.Flow):
        lines.extend([f"{c.SEPARATOR1} {i + 1}/{len(data1.Flow)} Flow"])
        lines.extend([f"{data1.Flow[i]}"])
        # Traces
        total = len(data1.Traces[i])
        for j, trace in enumerate(data1["Traces"][i]):
            separator = f"{c.SEPARATOR2} {j + 1}/{total} Trace"
            lines.extend(get_formatted_result(args, separator, trace))
    show_result_lines(args, "\n".join(lines))


def cmd_vlan(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Node", "VLAN_ID", "Interfaces", "VXLAN_VNI"]
    show1 = ["Node", "VLAN_ID", "Interfaces", "VXLAN_VNI"]
    data1 = (
        bf.q.switchedVlanProperties(
            nodes=get_query(args.node),
            excludeShutInterfaces=args.noshut,
        )
        .answer()
        .frame()
    )
    if args.vlan is not None:
        data1 = data1[data1["VLAN_ID"] == int(args.vlan)]
    show_result(args, data1[show1], sort1)


def cmd_vrrp(bf: Session, args: argparse.ArgumentParser):
    sort1 = ["Interface", "Group_Id"]
    show1 = [
        "Interface",
        "Group_Id",
        "Virtual_Addresses",
        "Source_Address",
        "Priority",
        "Preempt",
        "Active",
    ]
    data1 = bf.q.vrrpProperties(nodes=get_query(args.node)).answer().frame()
    if len(data1) != 0:
        data1["Virtual_Addresses"] = data1.apply(get_hsrp_vip, axis=1)
    if args.id is not None:
        data1 = data1[data1["Group_Id"] == args.id]
    if args.vip is not None:
        data1 = data1[data1["Virtual_Addresses"] == args.vip]
    if args.wopair:
        data1 = data1.drop_duplicates(
            subset=["Group_Id", "Virtual_Addresses"], keep=False
        )
    show_result(args, data1[show1], sort1)


def debug_bidir_flow(data: pandas.DataFrame, message: str):
    # Flow(s)
    try:
        flow_forward = data.Forward_Flow.count()
    except:
        flow_forward = "(None)"
    try:
        flow_reverse = data.Reverse_Flow.count()
    except:
        flow_reverse = "(None)"
    util.debug(
        f"{c.SEPARATOR2}",
        f"{message}",
        f" Forward Flow: {flow_forward}",
        f" Reverse Flow: {flow_reverse}",
    )

    # Trace(s))
    for i in range(len(data.Forward_Traces)):
        try:
            ftrace = len(data.Forward_Traces[i])
        except:
            ftrace = "(None)"
        try:
            rtrace = len(data.Reverse_Traces[i])
        except:
            rtrace = "(None)"
        util.debug(
            f"  Flow {i + 1} -> Forward Trace: {ftrace}",
            f"  Flow {i + 1} -> Reverse Trace: {rtrace}",
        )


def debug_flow(data: pandas.DataFrame, message: str):
    # Flow (s)
    try:
        flow = data.Flow.count()
    except:
        flow = "(None)"
    util.debug(f"{c.SEPARATOR2}", f"{message}", f" Flow: {flow}")
    # Trace(s))
    for i in range(len(data.Traces)):
        try:
            trace = len(data.Traces[i])
        except:
            trace = "(None)"
        util.debug(f"  Flow {i + 1} -> Trace: {trace}")


def get_bandwidth(row: pandas.Series) -> str:
    return calc_speed(row["Bandwidth"])


def get_formatted_result(args: argparse.ArgumentParser, separator: str, trace) -> list:
    lines = []
    if args.format == "all":
        lines = [separator, str(trace)]
    elif args.format == "node":
        lines = [separator, get_hop_nodes(trace)]
    return lines


def get_headers(args: argparse.ArgumentParser) -> HeaderConstraints:
    is_ack, is_cwr, is_ece, is_fin, is_psh, is_rst, is_syn, is_urg = (
        False,
        False,
        False,
        False,
        False,
        False,
        False,
        False,
    )
    if args.tcpflags is not None:
        if "ack" in args.tcpflags:
            is_ack = True
        if "cwr" in args.tcpflags:
            is_cwr = True
        if "ece" in args.tcpflags:
            is_ece = True
        if "fin" in args.tcpflags:
            is_fin = True
        if "psh" in args.tcpflags:
            is_psh = True
        if "rst" in args.tcpflags:
            is_rst = True
        if "syn" in args.tcpflags:
            is_syn = True
        if "urg" in args.tcpflags:
            is_urg = True
    if args.sport is None:
        args.sport = 49152
    if args.dport is None:
        args.dport = 33434
    header = HeaderConstraints(
        srcIps=args.saddr,
        dstIps=args.daddr,
        srcPorts=args.sport,
        dstPorts=args.dport,
        applications=args.application,
        ipProtocols=args.protocol,
        icmpCodes=args.icmpcodes,
        icmpTypes=args.icmptypes,
        dscps=args.dscps,
        ecns=args.ecns,
        packetLengths=args.lengths,
        fragmentOffsets=args.offsets,
        tcpFlags=MatchTcpFlags(
            TcpFlags(
                ack=is_ack,
                cwr=is_cwr,
                ece=is_ece,
                fin=is_fin,
                psh=is_psh,
                rst=is_rst,
                syn=is_syn,
                urg=is_urg,
            )
        ),
    )
    util.debug(
        "Packet Header:",
        f" Protocol        : {header.ipProtocols}",
        f" Application     : {header.applications}",
        f" ICMP Code       : {header.icmpCodes}",
        f" ICMP Type       : {header.icmpTypes}",
        f" Src IP          : {header.srcIps}",
        f" Src Port        : {header.srcPorts}",
        f" Dst IP          : {header.dstIps}",
        f" Dst Port        : {header.dstPorts}",
        f" DSCP            : {header.dscps}",
        f" ECN             : {header.ecns}",
        f" Packet Length   : {header.packetLengths}",
        f" Fragment Offset : {header.fragmentOffsets}",
        f" TCP Flags       : {header.tcpFlags}",
    )
    return header


def get_hop_nodes(trace) -> str:
    result = []
    nodes = str(trace).split("\n")
    for node in nodes:
        if not node.startswith(" "):
            result.append(node)
    return "\n".join(result)


def get_hsrp_vip(row: pandas.Series) -> str:
    return row["Virtual_Addresses"][0]


def get_os_version(path: str, keyword: str) -> str:
    with open(path) as f:
        for line in f:
            if line.startswith(keyword):
                return line[len(keyword) :].strip()
    return "?"


def get_query(values: list) -> str:
    if values is None:
        query = ".*"
    else:
        items = []
        for value in values:
            items.append("^" + value + "$")
        query = "/" + "|".join(items) + "/"
    return query


def get_speed(row: pandas.Series) -> str:
    return calc_speed(row["Speed"])


def get_start_location(args: argparse.ArgumentParser) -> str:
    if args.interface is None:
        start_location = '"' + args.node + '"'
    else:
        start_location = args.node + "[" + args.interface + "]"
    return start_location


def is_ipaddr(address: str):
    try:
        socket.inet_aton(address)
        return True
    except:
        return False


def reach(bf: Session, action: str, args: argparse.ArgumentParser) -> pandas.DataFrame:
    headers = get_headers(args)
    location = get_start_location(args)
    data = (
        bf.q.reachability(
            pathConstraints=PathConstraints(startLocation=location),
            headers=headers,
            actions=action.upper(),
        )
        .answer()
        .frame()
    )
    return data


def show_elapsed_time():
    time_end = time.perf_counter()
    elapsed_time = time_end - time_start
    util.debug_line()
    util.debug(f"Elapsed Time: {elapsed_time:.2f} sec.")


def set_style(cell: openpyxl.cell.cell.Cell, index: int) -> None:
    if index == 0:
        FG_COLOR = "FFFFFF"
        BG_COLOR = "000000"
        cell.fill = openpyxl.styles.PatternFill(
            patternType="solid", fgColor=BG_COLOR, bgColor=BG_COLOR
        )
    else:
        FG_COLOR = "000000"
        BG_COLOR = "FFFFFF"
    cell.font = openpyxl.styles.Font(color=FG_COLOR, name="MS Gothic")
    side = openpyxl.styles.Side(style="thin", color=FG_COLOR)
    cell.border = openpyxl.styles.Border(top=side, bottom=side, left=side, right=side)


def show_result(args: argparse.ArgumentParser, response: pandas.DataFrame, sort: list):
    if len(response) == 0:
        util.info("No data.")
        return
    else:
        print(response.sort_values(by=sort).to_string(index=False))
    if args.excel is None:
        return
    # Output to Excel.
    filename = args.excel
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    response.sort_values(by=sort).to_excel(filename, index=False)
    wb = openpyxl.load_workbook(filename)
    ws = wb["Sheet1"]
    ws.freeze_panes = "A2"
    # Window width adjustment
    for col in ws.iter_cols():
        max_length = 0
        column = col[0].column_letter
        for index, cell in enumerate(col):
            if cell.value is None:
                continue
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
            set_style(cell, index=index)
        ws.column_dimensions[column].width = (max_length + 2) * 1.2
    wb.save(filename)
    wb.close()
    util.info_line()
    util.info(f"Output: {filename}")


def show_result_lines(args: argparse.ArgumentParser, lines: list):
    lines = lines.splitlines()
    if args.excel is not None:
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        for line in lines:
            ws.append([line])
        wb.save(args.excel + ".xlsx")
        wb.close()
    else:
        for line in lines:
            print(line)
    show_elapsed_time()


def main():
    common = parser.common()
    detail = parser.detail()
    excel = parser.excel()
    filter = parser.filter()
    format = parser.format()
    header = parser.header()
    node = parser.node()
    rnode = parser.rnode()
    start = parser.start()
    summary = parser.summary()
    vrf = parser.vrf()

    # Main parser
    mainparser = argparse.ArgumentParser(
        prog=c.PROG_NAME,
        description=f"{c.PROG_NAME} ({c.PROG_DESCRIPTION}) {c.PROG_VERSION}",
        add_help=False,
    )
    subparsers = mainparser.add_subparsers(title="commands", metavar="{command}")

    # Search Filters
    _aclsearch = subparsers.add_parser(
        "aclsearch",
        help="Finds flows for which a filter takes a particular behavior.",
        parents=[common, excel, filter, header, node],
    )
    _aclsearch.set_defaults(handler=cmd_aclsearch)

    # Test Filters
    _acltest = subparsers.add_parser(
        "acltest",
        help="Returns how a flow is processed by a filter (ACLs, firewall rules).",
        parents=[common, excel, filter, header, node],
    )
    _acltest.set_defaults(handler=cmd_acltest)

    # Filter Line Reachability
    _aclunreach = subparsers.add_parser(
        "aclunreach",
        help="Returns unreachable lines in filters (ACLs and firewall rules).",
        parents=[common, excel, filter, node, summary],
    )
    _aclunreach.set_defaults(handler=cmd_aclunreach)

    # BGP Edges
    bgpedge = subparsers.add_parser(
        "bgpedge",
        help="Returns BGP adjacencies.",
        parents=[common, detail, excel, node],
    )
    bgpedge.set_defaults(handler=cmd_bgpedge)

    # BGP RIB
    bgprib = subparsers.add_parser(
        "bgprib",
        help="Returns routes in the BGP RIB.",
        parents=[common, detail, excel, node],
    )
    bgprib.set_defaults(handler=cmd_bgprib)

    # Bi-directional Reachability
    _bireach = subparsers.add_parser(
        "bireach",
        help="Searches for successfully delivered flows that can successfully receive a response.",
        parents=[common, excel, format, header, start],
    )
    _bireach.add_argument("--result", action="store_true", help="Display results only.")
    _bireach.add_argument(
        "--success", action="store_true", help="Search only SUCCESS flows."
    )
    _bireach.set_defaults(handler=cmd_bireach)

    # Bi-directional Traceroute
    _bitrace = subparsers.add_parser(
        "bitrace",
        help="Traces the path(s) for the specified flow, along with path(s) for reverse flows.",
        parents=[common, excel, format, header, start],
    )
    _bitrace.add_argument("--result", action="store_true", help="Display results only.")
    _bitrace.set_defaults(handler=cmd_bitrace)

    # EIGRP Edge
    _eigrpedge = subparsers.add_parser(
        "eigrpedge", help="EIGRP Edges", parents=[common, excel, node]
    )
    _eigrpedge.set_defaults(handler=cmd_eigrpedge)

    # HSRP
    _hsrp = subparsers.add_parser(
        "hsrp",
        help="Returns configuration settings of HSRP groups.",
        parents=[common, excel, node],
    )
    _hsrp.add_argument("--id", type=str, help="Only matching groups are displayed.")
    _hsrp.add_argument(
        "--vip",
        type=str,
        help="Include only groups with at least one virtual address matching this specifier.",
    )
    _hsrp.add_argument(
        "--wopair", action="store_true", help="Only unpaired settings are shown."
    )
    _hsrp.set_defaults(handler=cmd_hsrp)

    # Interface
    _interface = subparsers.add_parser(
        "interface",
        help="Returns configuration settings of interfaces.",
        parents=[common, excel, node],
    )
    _interface.add_argument(
        "--column",
        default="basic",
        choices=[
            "acl",
            "all",
            "basic",
            "hsrp",
            "ip",
            "portchannel",
            "speed",
            "state",
            "stp",
            "vlan",
        ],
        help="Select items to display. (Default: basic)",
    )
    _interface.add_argument(
        "--wounit",
        action="store_true",
        help="Do not display units.",
    )
    _interface.set_defaults(handler=cmd_interface)

    # IP
    _ip = subparsers.add_parser(
        "ip",
        help="Returns where IP addresses are attached in the network.",
        parents=[common, excel, node, vrf],
    )
    _ip.add_argument(
        "--address",
        default=".*",
        help="Restrict output to only specified IP addresses.",
    )
    _ip.add_argument(
        "--duplicate",
        action="store_true",
        help="Restrict output to only IP addresses that are duplicated (configured on a different node or VRF) in the snapshot.",
    )
    _ip.set_defaults(handler=cmd_ip)

    # Layer 3 Topology
    _l3edge = subparsers.add_parser(
        "l3edge",
        help="Lists all Layer 3 edges in the network.",
        parents=[common, excel, node, rnode],
    )
    _l3edge.set_defaults(handler=cmd_l3edge)

    # Loop detection
    _loop = subparsers.add_parser(
        "loop", help="Detects forwarding loops.", parents=[common, excel, format]
    )
    _loop.set_defaults(handler=cmd_loop)

    # Longest Prefix Match
    _lpm = subparsers.add_parser(
        "lpm", help="Returns routing tables.", parents=[common, excel, node]
    )
    _lpm.add_argument(
        "-a",
        "--address",
        required=True,
        help="Return routes for networks matching this prefix.",
    )
    _lpm.add_argument(
        "--vrf", help="Return routes on VRFs matching this name or regex.", nargs="*"
    )
    _lpm.set_defaults(handler=cmd_lpm)

    # Node Properties
    _node = subparsers.add_parser("node", parents=[common, excel, node])
    _node.add_argument(
        "--column",
        default="basic",
        choices=[
            "acl",
            "all",
            "basic",
            "bgp",
            "dns",
            "ipsec",
            "logging",
            "ntp",
            "route",
            "snmp",
            "tacacs",
        ],
        help="Select items to display. (Default: basic)",
    )
    _node.set_defaults(handler=cmd_node)

    # OSPF Edges
    _ospfedge = subparsers.add_parser(
        "ospfedge",
        help="Lists all OSPF adjacencies in the network.",
        parents=[common, excel, node],
    )
    _ospfedge.set_defaults(handler=cmd_ospfedge)

    # Parse
    _parse = subparsers.add_parser(
        "parse", help="Displays file parse status.", parents=[common, excel, node]
    )
    _parse.set_defaults(handler=cmd_parse)

    # Prefix
    _prefix = subparsers.add_parser(
        "prefix",
        help="Traces prefix propagation through the network.",
        parents=[common, excel, node],
    )
    _prefix.add_argument(
        "--prefix", help="The prefix to trace. Expected format is A.B.C.D/Y"
    )
    _prefix.set_defaults(handler=cmd_prefix)

    # Reach
    _reach = subparsers.add_parser(
        "reach",
        help="Finds flows that match the specified path and header space conditions.",
        parents=[common, excel, format, header, start],
    )
    _reach.set_defaults(handler=cmd_reach)

    # Route
    _route = subparsers.add_parser(
        "route",
        help="Returns routing tables.",
        parents=[common, detail, excel, node, vrf],
    )
    _route.add_argument(
        "--address", help="Return routes for networks matching this prefix."
    )
    _route.add_argument(
        "--match",
        default="longest",
        choices=["exact", "longest", "longer", "shorter"],
        help="Use this prefix matching criterion: exact(EXACT), longest(LONGEST_PREFIX_MATCH), longest(LONGER_PREFIXES), shorter(SHORTER_PREFIXES). (Default: longest)",
    )
    _route.add_argument(
        "--protocol",
        help="Return routes for protocols matching this specifier.",
        nargs="*",
    )
    _route.set_defaults(handler=cmd_route)

    # Trace
    _trace = subparsers.add_parser(
        "trace",
        help="Traces the path(s) for the specified flow.",
        parents=[common, excel, format, header, start],
    )
    _trace.set_defaults(handler=cmd_trace)

    # VLAN
    _vlan = subparsers.add_parser(
        "vlan",
        help="Returns configuration settings of switched VLANs.",
        parents=[common, excel, node],
    )
    _vlan.add_argument("--vlan", type=int, help="Include VLANs in this space.")
    _vlan.add_argument(
        "--noshut",
        action="store_true",
        help="Exclude interfaces that are shutdown.",
    )
    _vlan.set_defaults(handler=cmd_vlan)

    # VRRP
    _vrrp = subparsers.add_parser(
        "vrrp",
        help="Returns configuration settings of VRRP groups.",
        parents=[common, excel, node],
    )
    _vrrp.add_argument("--id", type=str, help="Only matching groups are displayed.")
    _vrrp.add_argument(
        "--vip",
        type=str,
        help="Include only groups with at least one virtual address matching this specifier.",
    )
    _vrrp.add_argument(
        "--wopair", action="store_true", help="Only unpaired settings are shown."
    )
    _vrrp.set_defaults(handler=cmd_vrrp)

    # If no subcommand is specified, help is displayed and the program exits.
    argcomplete.autocomplete(parser)
    args = mainparser.parse_args()
    if not hasattr(args, "handler"):
        mainparser.print_help()
        return

    # Set c.logger configuration.
    c.logger = logging.getLogger(__name__)
    handler = logging.StreamHandler(sys.stderr)
    handler.setFormatter(logging.Formatter("[%(levelname)s]%(message)s"))
    loglevels = {
        "DEBUG": logging.DEBUG,
        "INFO": logging.INFO,
        "WARNING": logging.WARNING,
        "ERROR": logging.ERROR,
        "CRITICAL": logging.CRITICAL,
    }
    c.logger.addHandler(handler)
    c.logger.setLevel(loglevels[args.log.upper()])
    util.debug_line()
    util.debug(f"{c.PROG_NAME} ({c.PROG_DESCRIPTION}) {c.PROG_VERSION}")
    util.debug_line()
    util.debug(f'Logging level set to "{args.log}".')

    # Display parameters
    util.debug_line()
    util.debug(
        "Batfish initialization parameters:",
        f" Host    : {args.batfish}",
        f" Port1   : {args.port1}",
        f" Port2   : {args.port2}",
        f" Timeout : {args.timeout}",
        f" HTTPS   : {args.https}",
        f" Insecure: {args.insecure}",
    )

    # Validate value (s)
    if not os.path.isdir(args.base):
        util.error("Base directory does not exist.")
        sys.exit(1)
    if not os.path.isdir(args.base + "/configs/"):
        util.error(f"'{args.base}/configs' directory does not exist.")
        sys.exit(1)

    pandas.options.display.max_colwidth = 100
    try:
        bf = util.Batfish(
            host=args.batfish,
            base=args.base,
            port1=args.port1,
            port2=args.port2,
            https=args.https,
            insecure=args.insecure,
            timeout=args.timeout,
            nocache=args.nocache,
        ).get_session()
    except Exception as e:
        util.error(f"{e}")
        sys.exit(1)
    util.debug_line()
    args.handler(bf, args)


if __name__ == "__main__":
    main()
