from decimal import Decimal
from pathlib import Path
from typing import List

import openpyxl

from sage_templater.plugin_manager import hookimpl
from sage_templater.schemas import SmallBoxRecordSchema


def get_wb_and_sheets(file_path: Path) -> (openpyxl.Workbook, List[str]):
    """Get workbook and sheets from an Excel file."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    return wb, wb.sheetnames


def get_start_and_end_row_numbers(wb: openpyxl.Workbook, sheet_name: str) -> tuple[int, int]:
    """Get start and end row numbers from a sheet with the small box format."""
    sheet = wb[sheet_name]
    start_row = -1
    end_row = sheet.max_row
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in ["Código", "CÓDIGO"]:
                start_row = cell.row
                break
    return start_row, end_row


def get_raw_rows(wb: openpyxl.Workbook, sheet_name: str, start_row: int, end_row: int) -> List[List[str]]:
    """Get raw rows from a sheet with the small box format."""
    sheet = wb[sheet_name]
    raw_rows = []
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        raw_row = []
        for cell in row:
            raw_row.append(str(cell.value))
        raw_rows.append(raw_row)
    return raw_rows


def parse_raw_rows(raw_rows: List[List[str]], source_file: Path, source_sheet: str) -> List[SmallBoxRecordSchema]:
    """Parse raw rows from a sheet with the small box format."""
    records = []
    for i, raw_row in enumerate(raw_rows, 1):
        if len(raw_row) < 10 or i == 1:
            continue
        if raw_row[6] is None or raw_row[6] == "None":
            break
        record = SmallBoxRecordSchema(
            code=raw_row[0],
            national_id=raw_row[1],
            verification_digit=raw_row[2],
            name=raw_row[3],
            invoice=raw_row[4],
            date=raw_row[5],
            amount=Decimal(raw_row[6]),
            tax=Decimal(raw_row[7]),
            total=Decimal(raw_row[8]),
            description=raw_row[9],
            source_file=str(source_file),
            source_sheet=source_sheet,
        )
        records.append(record)
    return records


@hookimpl
def parse_file(file_path: str) -> List[SmallBoxRecordSchema]:
    if not file_path.endswith(".xlsx"):
        return []
    return None
