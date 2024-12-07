"""
这个模块主要是用来处理excel文件相关的操作
"""
import openpyxl
import os
import time
import openpyxl.utils
import psutil
import pandas as pd


class ExcelUtils:
    """
    这个类主要用来处理excel文件相关的操作，包括读取、写入、合并单元格、获取单元格、获取合并单元格等功能。
    close_save_open_files: 设置为True时，保存操作或变为先强行关闭打开的文件（所有WPS或者Office进程），防止文件被占用，操作完成后再重新打开文件
    """

    def __init__(self, file_path, sheet_name='Sheet1'):
        if not os.path.exists(file_path) or not os.path.isfile(file_path):
            raise FileNotFoundError(f"文件 {file_path} 不存在或者不是文件")
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(self.file_path)
        self.worksheet = self.wb[sheet_name]
        self.max_row = self.worksheet.max_row
        self.max_column = self.worksheet.max_column
        self.cells_columns_iter = self.worksheet.columns
        self.merged_cells = self.worksheet.merged_cells
        self.merged_cells_ranges = self.merged_cells.ranges
        self.close_save_open_files = False

    @staticmethod
    def create_new_workbook(file_path, sheet_name='Sheet1'):
        """
        创建新的excel文件
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(file_path)
        print(f"创建新的excel文件 {file_path} 成功")

    def get_excel_column_index_by_name(self, *column_names):
        """
        获取包含指定字符串的列对应的索引，包括重复的列名
        """
        col_name_tuple = *column_names,
        col_index_dict = {}
        for col in self.cells_columns_iter:  # 迭代器，每次得到一列的单元格信息（第一行到表格整体最大行数）
            for cell in col:
                if cell.value in col_name_tuple:
                    col_index_dict[cell.value] = cell.column
        not_found_col_names = set(col_name_tuple) - set(col_index_dict.keys())
        # print(self.cells_columns_iter)
        self.cells_columns_iter = self.worksheet.columns
        print(f"未找到列名：{not_found_col_names}")
        return col_index_dict

    def convert_cells_format_o2c(self, row, column) -> tuple:
        """
        转换单元格格式，返回格式为 (coordinate, value)
        """
        return self.worksheet.cell(row=row, column=column).coordinate, self.worksheet.cell(row=row, column=column).value

    def convert_merge_cells_format_o2c(self, start_row, end_row, start_column, end_column) -> tuple:
        """
        转换合并单元格格式，返回格式为 (coordinate, value)
        """
        for merged_cell in self.merged_cells:
            if merged_cell.min_row == start_row and merged_cell.max_row == end_row and merged_cell.min_col == start_column and merged_cell.max_col == end_column:
                return merged_cell.coord, self.worksheet.cell(row=start_row, column=start_column).value
        print(f"未找到合并单元格：{start_row} - {end_row}, {start_column} - {end_column}")

    def convert_cells_format_c2o(self, coordinate) -> tuple:
        """
        转换单元格格式，返回格式为 (row, column, value)
        """
        row, column = openpyxl.utils.cell.coordinate_to_tuple(coordinate)
        return row, column, self.worksheet.cell(row=row, column=column).value

    def convert_merge_cells_format_c2o(self, coordinate1, coordinate2) -> tuple:
        """
        转换合并单元格格式，返回格式为 (start_row, end_row, start_column, end_column, value)
        """
        start_row, start_column = openpyxl.utils.cell.coordinate_to_tuple(coordinate1)
        end_row, end_column = openpyxl.utils.cell.coordinate_to_tuple(coordinate2)
        return start_row, end_row, start_column, end_column, self.worksheet.cell(row=start_row,
                                                                                 column=start_column).value

    def get_merge_cells_index(self):
        raise NotImplementedError("该方法尚未实现")

    def find_merged_cells_in_table(self, start_row=1, end_row=None, start_column=1, end_column=None) -> list:
        """
        查找某一范围表格合并的单元格, 返回格式为 [(start_row, end_row, start_column, end_column, value),...]
        """
        if end_row is None:
            end_row = self.max_row
        if end_column is None:
            end_column = self.max_column

        merged_cell_ranges = []
        for merged_cell in self.merged_cells:
            min_row, max_row, min_col, max_col = merged_cell.min_row, merged_cell.max_row, merged_cell.min_col, merged_cell.max_col
            if min_row >= start_row and max_row <= end_row and min_col >= start_column and max_col <= end_column:
                merged_cell_ranges.append(
                    [min_row, max_row, min_col, max_col, self.worksheet.cell(row=min_row, column=min_col).value])
        return merged_cell_ranges

    def find_merged_cells_in_column(self, column) -> list:
        """
        查找某一列合并的单元格, 返回格式为 [(start_row, end_row, start_column, end_column, value),...]
        :param column:
        :return:
        """
        merged_cell_ranges = []
        for merged_cell in self.merged_cells:
            min_row, max_row, min_col, max_col = merged_cell.min_row, merged_cell.max_row, merged_cell.min_col, merged_cell.max_col
            if min_col <= column <= max_col:
                merged_cell_ranges.append(
                    (min_row, max_row, min_col, max_col, self.worksheet.cell(row=min_row, column=min_col).value))
        return merged_cell_ranges

    def find_merged_cells_in_row(self, row) -> list:
        """
        查找某一行合并的单元格, 返回格式为 [(start_row, end_row, start_column, end_column, value),...]
        :param row:
        :return:
        """
        merged_cell_ranges = []
        for merged_cell in self.merged_cells:
            min_row, max_row, min_col, max_col = merged_cell.min_row, merged_cell.max_row, merged_cell.min_col, merged_cell.max_col
            if min_row <= row <= max_row:
                merged_cell_ranges.append(
                    (min_row, max_row, min_col, max_col, self.worksheet.cell(row=min_row, column=min_col).value))
        return merged_cell_ranges

    def judge_cell_is_merged(self, row, column, mode='o'):
        """
        判断单元格是否是合并单元格,mode选择是以数字坐标输入还是以字母坐标输入,返回（），元组第一个bool值表示是否是合并单元格，第二个bool值表示是否是合并单元格的起始单元格
        """
        if mode != 'c':
            coord = self.convert_cells_format_o2c(row, column)[0]
        else:
            coord = f"{row}{column}"
        for merged_cell_ranges in self.merged_cells_ranges:
            if coord in merged_cell_ranges:
                is_start_cell = False
                #                 print(f"单元格 {coord} 是合并单元格")
                # print(f'{merged_cell_ranges.min_row = }')
                if row == merged_cell_ranges.min_row:
                    #                     print(f"单元格 {coord} 是合并单元格的起始单元格")
                    is_start_cell = True
                else:
                    print(f"单元格 {coord} 不是合并单元格的起始单元格")
                return True, is_start_cell
        #         print(f"单元格 {coord} 不是合并单元格")
        return False, False

    def set_row_values(self, row, start_column, end_column, values):
        """
        设置一行单元格的值，包括合并单元格的值
        """
        j = 0
        for i in range(start_column, end_column + 1):
            if all(self.judge_cell_is_merged(row, i)):
                # 如果是合并单元格，则设置合并单元格的值
                self.worksheet.cell(row=row, column=i, value=values[j])
                j += 1
            if not self.judge_cell_is_merged(row, i)[0]:
                self.worksheet.cell(row=row, column=i, value=values[j])
                j += 1
        self.__close_save_open_files()

    def get_row_values(self, row, start_column=1, end_column=None, ignore_empty_cell=False) -> list:
        """
        获取一行单元格的值，忽略空单元格时就相当于包括了合并单元格的值
        """
        if end_column is None:
            end_column = self.max_column
        values = []
        for i in range(start_column, end_column + 1):
            if ignore_empty_cell and not ((not self.worksheet.cell(row=row, column=i).value) or (
                    "Unnamed" in self.worksheet.cell(row=row, column=i).value)):
                continue
            values.append(self.worksheet.cell(row=row, column=i).value)
        return values

    def set_column_values(self, column, start_row, end_row, values, ignore_empty_item=False):
        """
        设置一列单元格的值，包括合并单元格的值
        """
        max_j = len(values) - 1
        j = 0
        break_token = False
        for i in range(start_row, end_row + 1):
            if break_token:
                break
            if all(self.judge_cell_is_merged(i, column)):
                # 如果是合并单元格，则设置合并单元格的值
                while True:
                    if j > max_j:
                        break_token = True
                        break
                    if ignore_empty_item and not values[j]:
                        j += 1
                        continue
                    try:
                        self.worksheet.cell(row=i, column=column, value=values[j])
                        break
                    except:
                        j += 1
                j += 1
            if not self.judge_cell_is_merged(i, column)[0]:
                while True:
                    if j > max_j:
                        break_token = True
                        break
                    if ignore_empty_item and not values[j]:
                        j += 1
                        continue
                    try:
                        self.worksheet.cell(row=i, column=column, value=values[j])
                        # print(values[j])
                        break
                    except:
                        j += 1
                j += 1
        # print(111)
        self.__close_save_open_files()

    def get_column_values(self, col_index, ignore_empty_cell=False):
        """
        获取指定列的单元格值
        """
        col_values = []
        # print(col_index)
        for cells_column in self.worksheet.iter_cols(min_col=col_index, max_col=col_index, min_row=1,
                                                     max_row=self.max_row):
            # print(cells_column)
            for cell in cells_column:
                if ignore_empty_cell and ((not cell.value) or (cell.value == "Unnamed")):
                    continue
                # print(cell.value)
                col_values.append(cell.value)
        return col_values

    def add_empty_row(self, row, row_num=1):
        """
        在指定行前面添加空行，返回新增行的行号
        """
        self.worksheet.insert_rows(row, amount=row_num)
        print(f"在 {row} 行前面添加 {row_num} 行空行成功")
        self.__close_save_open_files()
        return [row + i for i in range(row_num)]

    def add_empty_column(self, col_index, col_num=1):
        """
        在指定列前面添加空列, 返回新增列的列号
        """
        self.worksheet.insert_cols(col_index, amount=col_num)
        print(f"在 {col_index} 列前面添加 {col_num} 列空列成功")
        self.__close_save_open_files()
        return [col_index + i for i in range(col_num)]

    def delete_rows(self, rows):
        """
        删除指定行
        """
        self.worksheet.delete_rows(rows)
        print(f"删除行 {rows} 成功")
        self.__close_save_open_files()

    def delete_columns(self, columns):
        """
        删除指定列
        """
        self.worksheet.delete_cols(columns)
        print(f"删除列 {columns} 成功")
        self.__close_save_open_files()

    def set_cells_format_to_string(self):
        """
        设置单元格格式为文本类型
        """
        for row in self.worksheet.rows:
            for cell in row:
                cell.number_format = '@'
        self.__close_save_open_files()

    def beatify_cells_width(self):
        """
        格式化单元格
        """
        for column in self.cells_columns_iter:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            # print(f'{column_letter = }')

            for cell in column:
                if cell.value:
                    # 如果有换行，依据换行符分割，取最大长度
                    if '\n' in str(cell.value):
                        max_length = max(max_length, max([len(line) for line in str(cell.value).split('\n')]))
                    else:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_width = 1.95 * max_length
            # print(f'{adjusted_width = }')
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
        self.cells_columns_iter = self.worksheet.columns
        self.__close_save_open_files()

    def beautify_cells_height(self):
        """
        格式化根据一行的单元格内容的最大行数来设置单元格高度
        """
        for row in self.worksheet.rows:
            max_height = 0
            for cell in row:
                if cell.value:
                    # 找\n出现了几次，则高度为\n的次数+1
                    height = str(cell.value).count('\n') + 1
                    max_height = max(max_height, height)
            adjusted_height = 15 * max_height
            # print(adjusted_height)
            self.worksheet.row_dimensions[row[0].row].height = adjusted_height
        self.__close_save_open_files()

    def __close_save_open_files(self):
        """
        关闭占用特定文件的进程。如果修改没能成功，可以考虑增大时间间隔
        """
        _time = 0.5
        try:
            self.wb.save(self.file_path)
        except PermissionError:
            if self.close_save_open_files:
                # 遍历所有进程

                for proc in psutil.process_iter(['pid', 'name']):
                    try:
                        # 检查进程名称是否为 'excel.exe'，不区分大小写
                        if proc.name().lower() in ['excel.exe', 'wps.exe']:
                            # 检查进程是否打开了指定的文件
                            print(f"关闭进程 {proc.name()} (PID: {proc.pid})，因为它 {self.file_path} 被打开")
                            # 尝试安全地终止进程
                            proc.terminate()
                            # 等待进程终止，设置超时时间
                            proc.wait(timeout=5)  # 超时时间可以根据需要调整
                            print(f"进程 {proc.name()} (PID: {proc.pid}) 已关闭")
                    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess) as e:
                        # 打印异常信息，但继续遍历其他进程
                        print(f"无法处理进程 {proc.pid}：{e},请考虑以管理员权限运行程序")
                    except psutil.TimeoutExpired:
                        # 如果进程没有在超时时间内终止，打印警告
                        print(f"进程 {proc.name()} (PID: {proc.pid}) 未能在规定时间内关闭，可能需要手动干预")
            time.sleep(1)  # 等待1秒，防止文件保存失败
            self.wb.save(self.file_path)

            os.startfile(self.file_path)

    def extract_non_empty_from_excel(self):
        # 从 Excel 文件中加载数据
        df = pd.read_excel(self.file_path)

        result_list = []

        # 遍历每一列，提取不为空的内容到结果列表
        for col in df.columns:
            col_data = df[col].dropna().tolist()
            result_list += col_data

        return result_list


if __name__ == '__main__':
    file_path = 'D:/test.xlsx'
    excel_utils = ExcelUtils(file_path)
    dic = excel_utils.get_excel_column_index_by_name('姓名', '年龄', '性别')
    print(dic)
    print(excel_utils.get_column_values(1, ignore_empty_cell=True))
    print(excel_utils.get_row_values(1, ignore_empty_cell=True))
    print(excel_utils.find_merged_cells_in_table())
    print(excel_utils.merged_cells_ranges)
    print(excel_utils.judge_cell_is_merged('D', '7', 'c'))
    print(excel_utils.convert_merge_cells_format_o2c(1, 1, 1, 2))
    excel_utils.set_row_values(5, 1, 3, ['afhdaouiefgaopiuefhgopwiuegf', 'b', 'c'])
    excel_utils.beatify_cells_width()
