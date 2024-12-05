import asyncio
from playwright.async_api import Playwright, async_playwright
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook
import os

def save(today_directory):
    filename = f"Bluelight_extracted_data_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    bluelight_filename = os.path.join(today_directory,filename)
    
    async def run(playwright: Playwright) -> None:
        # Playwright Part
        browser = await playwright.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()
        await page.goto("https://uk.eu-supply.com/ctm/supplier/publictenders?B=BLUELIGHT")

        today = datetime.now()
        to_date = today

        if to_date.weekday() == 0:
            from_date = to_date - timedelta(days=3)
        else:
            from_date = to_date - timedelta(days=1)

        to_date_str = to_date.strftime("%d/%m/%Y")
        from_date_str = from_date.strftime("%d/%m/%Y")

        print(f"From Date: {from_date_str}, To Date: {to_date_str}")

        await page.fill("#SearchFilter_FromDate", from_date_str)
        await page.fill("#SearchFilter_ToDate", to_date_str)

        await page.get_by_text("Search", exact=True).click()
        await page.wait_for_load_state("networkidle")
        
        workbook = Workbook()
        sheet = workbook.active
        headers = ["Title", "Link", "Other Info"]  # Update based on your table structure
        sheet.append(headers)

        # Function to scrape data from the current page
        async def scrape_data():
            soup = BeautifulSoup(await page.content(), "html.parser")
            table = soup.find('table')
            rows = []

            for row in table.find_all('tr')[1:]:
                cells = row.find_all(['td', 'th'])
                row_data = []
                for cell in cells:
                    anchor = cell.find('a')
                    if anchor:  # If there's an anchor tag
                        link = anchor['href']
                        if not link.startswith('http'):
                            link = "https://uk.eu-supply.com" + link
                        # Store only the name as hyperlink
                        row_data.append((anchor.get_text().strip(), link))  # Store as tuple
                    else:
                        row_data.append(cell.get_text().strip())
                rows.append(row_data)

            # Write data to Excel with hyperlinks
            next_row = sheet.max_row + 1
            for row_data in rows:
                for col, cell_data in enumerate(row_data, 1):
                    if isinstance(cell_data, tuple):  # If it's a (text, link) tuple
                        text, link = cell_data
                        sheet.cell(row=next_row, column=col).value = text
                        sheet.cell(row=next_row, column=col).hyperlink = link
                        sheet.cell(row=next_row, column=col).font = Font(color="0000EE", underline="single")
                    else:
                        sheet.cell(row=next_row, column=col).value = cell_data
                next_row += 1

        # Scrape the first page
        await scrape_data()

        # Pagination
        while True:
            # Check for the next page button
            next_page_button = await page.query_selector("a.pager-action:not(.state-active)")
            if next_page_button:
                await next_page_button.click()
                await page.wait_for_load_state("networkidle")  # Wait for the next page to load
                await scrape_data()  # Scrape the next page
            else:
                break  # No more pages

        # Save the workbook
        workbook.save(bluelight_filename)
        print(f"Data successfully written to {bluelight_filename}.")

        await context.close()
        await browser.close()

    async def main() -> None:
        async with async_playwright() as playwright:
            await run(playwright)

    asyncio.run(main())
    return bluelight_filename